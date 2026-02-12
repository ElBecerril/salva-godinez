"""Consolidador de libros Excel: unir archivos o unir hojas."""

import os

from rich.console import Console
from rich.prompt import Prompt
from rich.panel import Panel
from rich import box

console = Console()


def _get_openpyxl():
    """Import lazy de openpyxl."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        console.print(
            "[bold red]openpyxl no esta instalado.[/bold red]\n"
            "[dim]Ejecuta: pip install openpyxl[/dim]"
        )
        return None


def _merge_files(paths: list[str], output: str) -> int:
    """Modo 1: Unir archivos — cada archivo aporta sus hojas al workbook final.

    Returns:
        Numero total de hojas copiadas.
    """
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return 0

    from copy import copy

    dest_wb = openpyxl.Workbook()
    dest_wb.remove(dest_wb.active)
    total_sheets = 0

    for path in paths:
        src_wb = openpyxl.load_workbook(path, data_only=True)
        base = os.path.splitext(os.path.basename(path))[0]

        for ws in src_wb.worksheets:
            # Nombre unico para la hoja
            sheet_name = f"{base}_{ws.title}"[:31]
            dest_ws = dest_wb.create_sheet(title=sheet_name)

            for row in ws.iter_rows():
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        dest_cell.font = copy(cell.font)
                        dest_cell.fill = copy(cell.fill)
                        dest_cell.number_format = cell.number_format
                        dest_cell.alignment = copy(cell.alignment)

            # Copiar anchos de columna
            for col_letter, dim in ws.column_dimensions.items():
                dest_ws.column_dimensions[col_letter].width = dim.width

            total_sheets += 1

    dest_wb.save(output)
    return total_sheets


def _merge_sheets(filepath: str, output: str) -> int:
    """Modo 2: Unir hojas — stack vertical de hojas en una sola.

    Returns:
        Numero total de filas en la hoja final.
    """
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return 0

    src_wb = openpyxl.load_workbook(filepath, data_only=True)
    dest_wb = openpyxl.Workbook()
    dest_ws = dest_wb.active
    dest_ws.title = "Consolidado"

    current_row = 1
    for idx, ws in enumerate(src_wb.worksheets):
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Skip header en hojas 2+ (fila 1)
            if idx > 0 and row_num == 1:
                continue
            for col, value in enumerate(row, 1):
                dest_ws.cell(row=current_row, column=col, value=value)
            current_row += 1

    dest_wb.save(output)
    return current_row - 1


def consolidator_menu() -> None:
    """Menu del consolidador de libros Excel."""
    console.print(
        Panel(
            "[bold]1[/bold] - Unir archivos (varios .xlsx → un workbook)\n"
            "[bold]2[/bold] - Unir hojas (hojas de un archivo → una sola hoja)\n"
            "[bold]0[/bold] - Volver",
            title="[bold yellow]Consolidador de Libros Excel[/bold yellow]",
            box=box.ROUNDED,
        )
    )
    mode = Prompt.ask("[bold cyan]Opcion[/bold cyan]", default="0")

    if mode == "1":
        console.print("\n[bold cyan]Unir archivos Excel[/bold cyan]")
        console.print("[dim]Ingresa las rutas de los archivos (vacio para terminar):[/dim]\n")

        paths = []
        while True:
            path = Prompt.ask(
                f"  [bold]Archivo {len(paths) + 1}[/bold] (vacio para terminar)",
                default="",
            ).strip().strip('"')
            if not path:
                break
            if not os.path.isfile(path):
                console.print(f"  [red]Archivo no encontrado: {path}[/red]")
                continue
            if not path.lower().endswith((".xlsx", ".xlsm")):
                console.print("  [red]Solo se soportan archivos .xlsx y .xlsm[/red]")
                continue
            paths.append(path)

        if len(paths) < 2:
            console.print("[yellow]Se necesitan al menos 2 archivos.[/yellow]")
            return

        output = Prompt.ask(
            "[bold]Ruta del archivo de salida[/bold]",
            default=os.path.join(os.path.dirname(paths[0]), "consolidado.xlsx"),
        ).strip().strip('"')

        with console.status("[bold green]Consolidando archivos..."):
            sheets = _merge_files(paths, output)

        if sheets:
            console.print(f"\n[bold green]Consolidado creado: {output} ({sheets} hojas)[/bold green]")

    elif mode == "2":
        console.print("\n[bold cyan]Unir hojas en una sola[/bold cyan]\n")

        filepath = Prompt.ask("[bold]Ruta del archivo Excel[/bold]").strip().strip('"')
        if not os.path.isfile(filepath):
            console.print("[red]Archivo no encontrado.[/red]")
            return

        base, ext = os.path.splitext(filepath)
        output = Prompt.ask(
            "[bold]Ruta del archivo de salida[/bold]",
            default=f"{base}_consolidado{ext}",
        ).strip().strip('"')

        with console.status("[bold green]Uniendo hojas..."):
            rows = _merge_sheets(filepath, output)

        if rows:
            console.print(f"\n[bold green]Archivo creado: {output} ({rows} filas)[/bold green]")

    elif mode == "0":
        return
    else:
        console.print("[red]Opcion no valida.[/red]")
