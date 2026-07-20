"""Comparador de Excel: diferencias celda por celda entre dos archivos."""

import os
import zipfile

from rich.prompt import Prompt
from rich.table import Table

from utils import get_openpyxl as _get_openpyxl, console


# --- Logica (sin UI) ---


def _safe_output_path(path: str) -> str:
    """Evita sobrescribir un archivo existente agregando un sufijo numerico."""
    if not os.path.exists(path):
        return path

    base, ext = os.path.splitext(path)
    counter = 1
    while True:
        candidate = f"{base}_{counter}{ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def compare_sheets(ws1, ws2) -> list[dict]:
    """Compara dos hojas celda por celda.

    Returns:
        Lista de diffs: [{celda, v1, v2}]
    """
    diffs = []
    max_row = max(ws1.max_row or 1, ws2.max_row or 1)
    max_col = max(ws1.max_column or 1, ws2.max_column or 1)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            v1 = ws1.cell(row=row, column=col).value
            v2 = ws2.cell(row=row, column=col).value
            if v1 != v2:
                cell_ref = ws1.cell(row=row, column=col).coordinate
                diffs.append({
                    "celda": cell_ref,
                    "v1": str(v1) if v1 is not None else "(vacio)",
                    "v2": str(v2) if v2 is not None else "(vacio)",
                })
    return diffs


def compare_files(path1: str, path2: str) -> dict:
    """Compara dos archivos Excel.

    Returns:
        dict con ok, error, sheets_only_v1, sheets_only_v2, common_diffs {sheet: [diffs]}
    """
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return {"ok": False, "error": None}

    wb1 = None
    wb2 = None
    try:
        wb1 = openpyxl.load_workbook(path1, data_only=True)
        wb2 = openpyxl.load_workbook(path2, data_only=True)
    except (
        openpyxl.utils.exceptions.InvalidFileException,
        zipfile.BadZipFile,
        KeyError,
        OSError,
    ) as e:
        if wb1 is not None:
            try:
                wb1.close()
            except Exception:
                pass
        return {"ok": False, "error": str(e)}

    names1 = set(wb1.sheetnames)
    names2 = set(wb2.sheetnames)

    result = {
        "ok": True,
        "error": None,
        "sheets_only_v1": sorted(names1 - names2),
        "sheets_only_v2": sorted(names2 - names1),
        "common_diffs": {},
        "_wb1": wb1,
        "_wb2": wb2,
    }

    for name in sorted(names1 & names2):
        diffs = compare_sheets(wb1[name], wb2[name])
        if diffs:
            result["common_diffs"][name] = diffs

    return result


def _generate_diff_report(path1: str, path2: str, comparison: dict, output: str) -> None:
    """Genera un reporte .xlsx con las diferencias marcadas en rojo."""
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return

    from openpyxl.styles import PatternFill

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    wb1 = comparison["_wb1"]
    # Marcar las celdas diferentes en rojo
    for sheet_name, diffs in comparison["common_diffs"].items():
        ws = wb1[sheet_name]
        for diff in diffs:
            cell = ws[diff["celda"]]
            cell.fill = red_fill

    wb1.save(output)
    wb1.close()


# --- Interfaz de consola ---


def comparator_menu() -> None:
    """Menu del comparador de Excel."""
    console.print("\n[bold cyan]Comparar dos Archivos de Excel[/bold cyan]\n")

    path1 = Prompt.ask("[bold]Ruta del primer archivo[/bold]").strip().strip('"')
    if not os.path.isfile(path1):
        console.print("[red]Archivo no encontrado.[/red]")
        return

    path2 = Prompt.ask("[bold]Ruta del segundo archivo[/bold]").strip().strip('"')
    if not os.path.isfile(path2):
        console.print("[red]Archivo no encontrado.[/red]")
        return

    with console.status("[bold green]Comparando archivos..."):
        result = compare_files(path1, path2)

    if not result.get("ok"):
        if result.get("error"):
            console.print(
                f"[red]No se pudo abrir uno de los archivos Excel "
                f"(corrupto, extension incorrecta o formato invalido): {result['error']}[/red]"
            )
        return

    wb1 = result.pop("_wb1", None)
    wb2 = result.pop("_wb2", None)

    try:
        # Hojas exclusivas
        if result["sheets_only_v1"]:
            console.print(f"\n[yellow]Hojas solo en archivo 1:[/yellow] {', '.join(result['sheets_only_v1'])}")
        if result["sheets_only_v2"]:
            console.print(f"[yellow]Hojas solo en archivo 2:[/yellow] {', '.join(result['sheets_only_v2'])}")

        # Diferencias
        total_diffs = sum(len(d) for d in result["common_diffs"].values())

        if total_diffs == 0:
            console.print("\n[bold green]Los archivos son identicos en las hojas comunes.[/bold green]")
            return

        console.print(f"\n[bold yellow]{total_diffs} diferencia(s) encontrada(s):[/bold yellow]")

        for sheet_name, diffs in result["common_diffs"].items():
            table = Table(title=f"Hoja: {sheet_name}")
            table.add_column("Celda", style="bold cyan")
            table.add_column("Archivo 1", style="red")
            table.add_column("Archivo 2", style="green")

            for diff in diffs[:30]:
                table.add_row(diff["celda"], diff["v1"], diff["v2"])

            if len(diffs) > 30:
                table.add_row("...", f"+{len(diffs) - 30} mas", "")

            console.print(table)

        # Ofrecer generar reporte
        gen_report = Prompt.ask(
            "\n[bold]Generar reporte con diferencias marcadas?[/bold]",
            choices=["s", "n"], default="s",
        )
        if gen_report == "s":
            base, ext = os.path.splitext(path1)
            output = f"{base}_comparado{ext}"
            output = Prompt.ask("[bold]Ruta del reporte[/bold]", default=output).strip().strip('"')

            # El reporte reutiliza wb1 (cargado con data_only=True, que
            # reemplaza formulas por sus valores). Si se guardara sobre
            # path1 o path2, se destruiria el archivo original.
            if os.path.abspath(output) in (os.path.abspath(path1), os.path.abspath(path2)):
                console.print(
                    "[red]La ruta del reporte no puede ser igual a ninguno de los "
                    "archivos originales (se perderian formulas/datos). Elige otra ruta.[/red]"
                )
                return
            output = _safe_output_path(output)
            result["_wb1"] = wb1
            _generate_diff_report(path1, path2, result, output)
            console.print(f"\n[bold green]Reporte guardado: {output}[/bold green]")
    finally:
        if wb1:
            wb1.close()
        if wb2:
            wb2.close()
